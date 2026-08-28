import os
import openpyxl
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Pattern Analyzer", layout="wide")
st.title("📊 Pattern Analysis Report")

# 1. फ़ाइल पाथ सेट करना (GitHub Root Folder या File Uploader)
base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
default_file_path = os.path.join(base_path, "Colour combination file.xlsx")

wb = None

# स्क्रीन पर फ़ाइल अपलोड का ऑप्शन (ताकि एरर कभी न आए)
uploaded_file = st.file_uploader("एक्सेल फ़ाइल चुनें (या डिफ़ॉल्ट फ़ाइल उपयोग करें)", type=["xlsx"])

if uploaded_file is not None:
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    st.success("अपलोड की गई फ़ाइल से डेटा लोड हो गया है!")
elif os.path.exists(default_file_path):
    wb = openpyxl.load_workbook(default_file_path, data_only=True)
    st.info("सिस्टम फ़ाइल ('Colour combination file.xlsx') से डेटा लोड हो गया है।")
else:
    st.error("कृपया अपनी 'Colour combination file.xlsx' फ़ाइल ऊपर अपलोड करें।")

# 2. एनालिसिस शुरू करना
if wb is not None:
    report_data = []

    if 'Sheet9' in wb.sheetnames:
        ws = wb['Sheet9']
        for r in range(4, 35):  # 1 से 31 तारीख तक स्कैन
            tariq = ws.cell(row=r, column=7).value
            if not tariq:
                continue
            
            # 2026, 2025, और 2024 के नंबर निकालना
            v26 = [str(ws.cell(row=r, column=c).value).strip() for c in [8, 9, 11, 12, 13, 14] if ws.cell(row=r, column=c).value and str(ws.cell(row=r, column=c).value).strip() != 'XX']
            v25 = [str(ws.cell(row=r, column=c).value).strip() for c in [18, 19, 21, 22, 23, 24] if ws.cell(row=r, column=c).value and str(ws.cell(row=r, column=c).value).strip() != 'XX']
            v24 = [str(ws.cell(row=r, column=c).value).strip() for c in [27, 28, 30, 31, 32, 33] if ws.cell(row=r, column=c).value and str(ws.cell(row=r, column=c).value).strip() != 'XX']
            
            # रिपीट नंबर खोजना
            match_25 = set(v26).intersection(set(v25))
            match_24 = set(v26).intersection(set(v24))
            
            report_data.append({
                "तारीख (Date)": f"तारीख {tariq}",
                "जून 2026": ", ".join(v26),
                "जून 2025": ", ".join(v25),
                "2026 vs 2025 रिपीट": ", ".join(match_25) if match_25 else "None",
                "2026 vs 2024 रिपीट": ", ".join(match_24) if match_24 else "None"
            })

        # 3. परिणाम दिखाना
        if report_data:
            df = pd.DataFrame(report_data)
            st.dataframe(df, use_container_width=True)
            
            # एक्सेल डाउनलोड बटन
            csv_data = df.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="📥 रिपोर्ट CSV में डाउनलोड करें",
                data=csv_data,
                file_name="Pattern_Analysis_Report.csv",
                mime="text/csv"
            )
    else:
        st.warning("फ़ाइल में 'Sheet9' नहीं मिली।")
